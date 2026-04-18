# =============================================================================
# 02_automatizar_excel_csv.py — Automatización de CSV y Excel
# =============================================================================
# CSV y Excel son los formatos de datos más usados en entornos empresariales.
# Python permite procesarlos, transformarlos y generar reportes automáticamente.
#
# Módulos:
#   csv      — incluido con Python, para leer/escribir CSV
#   openpyxl — para archivos Excel .xlsx (pip install openpyxl)
#
# Instalación:
#   pip install openpyxl
#
# Contenido:
#   - Leer y escribir CSV con el módulo csv nativo
#   - Transformar y filtrar datos CSV
#   - Consolidar múltiples CSVs en uno
#   - Crear y leer archivos Excel con openpyxl
#   - Generar reportes con formato (colores, estilos, gráficos)
# =============================================================================

import csv
import os
import tempfile
import shutil
from pathlib import Path
from datetime import datetime, date
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# =============================================================================
# DATOS DE PRUEBA
# =============================================================================

# Datos de ventas para los ejemplos
SALES_DATA_ENERO = [
    ["fecha",       "vendedor",    "producto",       "cantidad", "precio_unitario"],
    ["2024-01-05",  "Ana García",  "Laptop Pro",     2,          1299.99],
    ["2024-01-08",  "Carlos López","Mouse Inalámbrico",10,       29.99],
    ["2024-01-12",  "Ana García",  "Teclado Mecánico",3,         89.99],
    ["2024-01-15",  "María Torres","Monitor 4K",     1,          399.99],
    ["2024-01-20",  "Carlos López","Laptop Pro",     1,          1299.99],
    ["2024-01-25",  "María Torres","Mouse Inalámbrico",5,        29.99],
    ["2024-01-28",  "Ana García",  "Monitor 4K",     2,          399.99],
]

SALES_DATA_FEBRERO = [
    ["fecha",       "vendedor",    "producto",       "cantidad", "precio_unitario"],
    ["2024-02-03",  "Ana García",  "Laptop Pro",     1,          1299.99],
    ["2024-02-07",  "Pedro Ruiz",  "Teclado Mecánico",4,         89.99],
    ["2024-02-10",  "Carlos López","Monitor 4K",     2,          399.99],
    ["2024-02-14",  "María Torres","Laptop Pro",     3,          1299.99],
    ["2024-02-18",  "Pedro Ruiz",  "Mouse Inalámbrico",8,        29.99],
    ["2024-02-22",  "Ana García",  "Teclado Mecánico",2,         89.99],
    ["2024-02-28",  "Carlos López","Laptop Pro",     2,          1299.99],
]


# =============================================================================
# OPERACIONES CON CSV
# =============================================================================

def write_csv(filepath, data):
    """
    Escribe datos en un archivo CSV.

    Parámetros:
        filepath (Path | str): ruta del archivo a crear
        data (list): lista de listas — primera fila son los encabezados

    Nota: newline="" es obligatorio en Windows para evitar líneas en blanco extra.
    encoding="utf-8-sig" añade BOM para que Excel abra el CSV con tildes correctamente.
    """
    with open(filepath, "w", newline="", encoding="utf-8-sig") as csvfile:
        writer = csv.writer(csvfile)
        writer.writerows(data)


def read_csv_as_dicts(filepath):
    """
    Lee un CSV y devuelve los datos como lista de diccionarios.

    DictReader usa la primera fila como claves del diccionario,
    lo que hace el código más legible: fila["vendedor"] vs fila[1].

    Parámetros:
        filepath (Path | str): ruta del CSV a leer

    Retorna:
        list[dict]: lista de diccionarios, uno por fila de datos
    """
    rows = []
    with open(filepath, "r", encoding="utf-8-sig") as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            rows.append(dict(row))
    return rows


def demo_csv_basico(work_dir):
    """
    Demuestra lectura y escritura básica de CSV.

    Parámetros:
        work_dir (Path): carpeta de trabajo para los archivos
    """
    print("\n--- CSV Básico: Leer y Escribir ---")

    # Crear CSV de enero
    enero_path = work_dir / "ventas_enero.csv"
    write_csv(enero_path, SALES_DATA_ENERO)
    print(f"  CSV creado: {enero_path.name}")

    # Leer como lista de diccionarios
    rows = read_csv_as_dicts(enero_path)
    print(f"  Filas leídas: {len(rows)}")
    print(f"  Columnas: {list(rows[0].keys())}")

    # Mostrar primeras 3 filas
    print("\n  Primeras 3 ventas:")
    for row in rows[:3]:
        total = float(row["cantidad"]) * float(row["precio_unitario"])
        print(f"    {row['fecha']} | {row['vendedor']:<15} | {row['producto']:<20} | ${total:,.2f}")


def filter_and_transform_csv(filepath, output_path, min_total=200):
    """
    Filtra filas de un CSV según un criterio y añade columnas calculadas.

    Parámetros:
        filepath (Path): CSV de entrada
        output_path (Path): CSV de salida (con datos transformados)
        min_total (float): valor mínimo de venta para incluir la fila

    Retorna:
        int: número de filas que pasaron el filtro
    """
    rows = read_csv_as_dicts(filepath)

    # Transformar: añadir columna 'total' y filtrar por mínimo
    filtered = []
    for row in rows:
        cantidad = float(row["cantidad"])
        precio   = float(row["precio_unitario"])
        total    = cantidad * precio
        row["total"] = f"{total:.2f}"

        if total >= min_total:
            filtered.append(row)

    # Escribir CSV filtrado
    if filtered:
        fieldnames = list(filtered[0].keys())
        with open(output_path, "w", newline="", encoding="utf-8-sig") as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(filtered)

    return len(filtered)


def consolidate_csvs(csv_files, output_path):
    """
    Consolida múltiples archivos CSV con la misma estructura en uno solo.

    Caso de uso real: reportes mensuales de ventas, datos de distintas
    tiendas o sucursales, logs de múltiples servidores, etc.

    Parámetros:
        csv_files (list): lista de rutas de CSV a consolidar
        output_path (Path): ruta del CSV consolidado de salida

    Retorna:
        int: total de filas consolidadas
    """
    all_rows = []
    headers = None

    for csv_path in csv_files:
        rows = read_csv_as_dicts(csv_path)
        if not rows:
            continue

        # Los encabezados se toman del primer archivo
        if headers is None:
            headers = list(rows[0].keys())

        # Añadir columna 'archivo_origen' para saber de qué CSV vino cada fila
        source = Path(csv_path).stem
        for row in rows:
            row["archivo_origen"] = source
            all_rows.append(row)

    if all_rows:
        fieldnames = headers + ["archivo_origen"]
        with open(output_path, "w", newline="", encoding="utf-8-sig") as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(all_rows)

    return len(all_rows)


def generate_summary_report(consolidated_path, output_path):
    """
    Genera un reporte de resumen por vendedor a partir del CSV consolidado.

    Parámetros:
        consolidated_path (Path): CSV consolidado de entrada
        output_path (Path): CSV de resumen de salida
    """
    rows = read_csv_as_dicts(consolidated_path)

    # Acumular totales por vendedor usando defaultdict
    by_seller = defaultdict(lambda: {"ventas": 0, "total": 0.0, "productos": set()})

    for row in rows:
        vendedor = row["vendedor"]
        total    = float(row["cantidad"]) * float(row["precio_unitario"])
        by_seller[vendedor]["ventas"] += 1
        by_seller[vendedor]["total"]  += total
        by_seller[vendedor]["productos"].add(row["producto"])

    # Convertir a lista de diccionarios para escribir CSV
    summary_rows = []
    for vendedor, stats in sorted(by_seller.items(), key=lambda x: -x[1]["total"]):
        summary_rows.append({
            "vendedor":           vendedor,
            "num_ventas":         stats["ventas"],
            "total_generado":     f"{stats['total']:.2f}",
            "productos_distintos": len(stats["productos"]),
        })

    with open(output_path, "w", newline="", encoding="utf-8-sig") as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=list(summary_rows[0].keys()))
        writer.writeheader()
        writer.writerows(summary_rows)

    return summary_rows


def demo_csv_avanzado(work_dir):
    """
    Demuestra filtrado, transformación y consolidación de CSVs.

    Parámetros:
        work_dir (Path): carpeta de trabajo
    """
    print("\n--- CSV Avanzado: Transformar y Consolidar ---")

    # Crear los CSV de los dos meses
    enero_path   = work_dir / "ventas_enero.csv"
    febrero_path = work_dir / "ventas_febrero.csv"
    write_csv(enero_path,   SALES_DATA_ENERO)
    write_csv(febrero_path, SALES_DATA_FEBRERO)

    # Filtrar ventas importantes (total >= $200)
    filtrado_path = work_dir / "ventas_enero_filtrado.csv"
    count = filter_and_transform_csv(enero_path, filtrado_path, min_total=200)
    print(f"  Enero filtrado (total >= $200): {count} ventas guardadas")

    # Consolidar enero y febrero en un solo CSV
    consolidado_path = work_dir / "ventas_consolidado.csv"
    total = consolidate_csvs([enero_path, febrero_path], consolidado_path)
    print(f"  CSV consolidado: {total} filas totales de 2 meses")

    # Generar reporte de resumen por vendedor
    resumen_path = work_dir / "resumen_vendedores.csv"
    summary = generate_summary_report(consolidado_path, resumen_path)

    print("\n  Resumen por vendedor (enero + febrero):")
    print(f"  {'Vendedor':<18} {'Ventas':<8} {'Total':<14} {'Productos únicos'}")
    print("  " + "-" * 55)
    for row in summary:
        print(f"  {row['vendedor']:<18} {row['num_ventas']:<8} ${row['total_generado']:<13} {row['productos_distintos']}")


# =============================================================================
# EXCEL CON OPENPYXL
# =============================================================================

def demo_excel(work_dir):
    """
    Demuestra creación y lectura de archivos Excel con openpyxl.

    Parámetros:
        work_dir (Path): carpeta de trabajo
    """
    if not OPENPYXL_AVAILABLE:
        print("\n--- Excel con openpyxl ---")
        print("  openpyxl no está instalado.")
        print("  Para instalarlo: pip install openpyxl")
        return

    print("\n--- Excel con openpyxl ---")

    excel_path = work_dir / "reporte_ventas.xlsx"

    # Crear un Workbook (libro de Excel)
    wb = openpyxl.Workbook()

    # La hoja activa por defecto se llama "Sheet"
    ws = wb.active
    ws.title = "Ventas"

    # Estilos para el encabezado
    header_font   = Font(bold=True, color="FFFFFF", size=11)
    header_fill   = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
    center_align  = Alignment(horizontal="center", vertical="center")

    # Estilo para filas alternas (cebra)
    alt_fill = PatternFill(start_color="E8EEF4", end_color="E8EEF4", fill_type="solid")

    # Escribir encabezados con estilo
    headers = ["Fecha", "Vendedor", "Producto", "Cantidad", "Precio Unit.", "Total"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align

    # Combinar datos de ambos meses (sin la fila de encabezado)
    all_data = SALES_DATA_ENERO[1:] + SALES_DATA_FEBRERO[1:]

    # Escribir datos con filas alternas y totales calculados
    for row_idx, row_data in enumerate(all_data, start=2):
        fecha, vendedor, producto, cantidad, precio = row_data
        total = float(cantidad) * float(precio)

        ws.cell(row=row_idx, column=1, value=str(fecha))
        ws.cell(row=row_idx, column=2, value=vendedor)
        ws.cell(row=row_idx, column=3, value=producto)
        ws.cell(row=row_idx, column=4, value=int(cantidad))
        ws.cell(row=row_idx, column=5, value=float(precio))
        ws.cell(row=row_idx, column=6, value=round(total, 2))

        # Formato de moneda para precio y total
        ws.cell(row=row_idx, column=5).number_format = "$#,##0.00"
        ws.cell(row=row_idx, column=6).number_format = "$#,##0.00"

        # Fila alterna
        if row_idx % 2 == 0:
            for col in range(1, 7):
                ws.cell(row=row_idx, column=col).fill = alt_fill

    # Fila de totales al final
    last_row = len(all_data) + 2
    ws.cell(row=last_row, column=2, value="TOTAL GENERAL").font = Font(bold=True)

    total_formula = f"=SUM(F2:F{last_row - 1})"
    total_cell = ws.cell(row=last_row, column=6, value=total_formula)
    total_cell.font = Font(bold=True)
    total_cell.number_format = "$#,##0.00"

    # Ajustar ancho de columnas automáticamente
    column_widths = [12, 18, 22, 10, 14, 12]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Fijar la primera fila (encabezado siempre visible al hacer scroll)
    ws.freeze_panes = "A2"

    # Guardar el archivo Excel
    wb.save(excel_path)
    print(f"  Excel creado: {excel_path.name}")
    print(f"  Hoja 'Ventas': {len(all_data)} filas de datos + encabezado + total")

    # Leer el Excel que acabamos de crear (verificación)
    wb_check = openpyxl.load_workbook(excel_path)
    ws_check  = wb_check["Ventas"]

    print(f"\n  Verificación de lectura:")
    print(f"  Hojas: {wb_check.sheetnames}")
    print(f"  Dimensiones: {ws_check.dimensions}")
    print(f"  Total de filas (incluyendo header): {ws_check.max_row}")

    wb_check.close()


# =============================================================================
# PROGRAMA PRINCIPAL
# =============================================================================

def main():
    """Función principal que ejecuta todas las demostraciones de CSV y Excel."""

    print("=" * 60)
    print("  DEMO: Automatización de CSV y Excel con Python")
    print("=" * 60)

    # Crear carpeta temporal de trabajo
    work_dir = Path(tempfile.mkdtemp(prefix="python_excel_demo_"))
    print(f"\nCarpeta de trabajo: {work_dir}")

    try:
        demo_csv_basico(work_dir)
        demo_csv_avanzado(work_dir)
        demo_excel(work_dir)

    finally:
        # Limpiar archivos generados
        shutil.rmtree(work_dir, ignore_errors=True)
        print(f"\nCarpeta temporal eliminada.")

    print("\n" + "=" * 60)
    print("  Demo completado.")
    print("=" * 60)


if __name__ == "__main__":
    main()
